import os
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Экспорт результатов парсинга в Excel (.xlsx)."""

    def __init__(self, output_path: str = "results.xlsx"):
        self.output_path = output_path

    def export(self, data: List[Dict], sheet_name: str = "Парсинг") -> str:
        """
        data — список словарей с ключами: name, price, old_price, image, url, source
        Возвращает путь к созданному файлу.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = ["Название", "Цена", "Старая цена", "Изображение", "URL", "Источник"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_num, item in enumerate(data, 2):
            row_data = [
                item.get("name", ""),
                item.get("price", ""),
                item.get("old_price", "") or "",
                item.get("image", ""),
                item.get("url", ""),
                item.get("source", ""),
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if col_num == 2 and value:
                    cell.number_format = '#,##0 "₽"'

        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column].width = adjusted_width

        ws.freeze_panes = "A2"

        ws.auto_filter.ref = ws.dimensions

        wb.save(self.output_path)
        return os.path.abspath(self.output_path)